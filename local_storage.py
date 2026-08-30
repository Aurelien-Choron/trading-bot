"""
Stockage local — Log les transactions et le portefeuille en CSV + Excel.
Remplace google_sheets.py pour un usage local (sans dépendance réseau).
"""

import os
import csv
from datetime import date, datetime

import openpyxl
from openpyxl.utils import get_column_letter


# Dossier de stockage
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)

# Fichiers
CSV_TRANSACTIONS = os.path.join(DATA_DIR, "transactions.csv")
CSV_PORTFOLIO = os.path.join(DATA_DIR, "portefeuille.csv")
CSV_JOURNAL = os.path.join(DATA_DIR, "journal.csv")
EXCEL_FILE = os.path.join(DATA_DIR, "trading_bot.xlsx")

# Headers
HEADERS_TRANSACTIONS = [
    "Date", "Heure", "Action", "Ticker", "Parts", "Prix",
    "Montant (€)", "Frais (€)", "PnL (€)", "Cash restant (€)", "Raisonnement"
]
HEADERS_PORTFOLIO = [
    "Date", "Valeur totale (€)", "Cash (€)", "Positions (€)",
    "Total investi (€)", "PnL (€)", "PnL (%)", "Nb positions",
    "Frais cumulés (€)", "Sentiment IA", "Niveau risque"
]
HEADERS_JOURNAL = [
    "Date", "Analyse marché", "Sentiment", "Risque",
    "Nb transactions", "Détail décisions"
]


def _append_csv(filepath: str, headers: list[str], row: list):
    """Ajoute une ligne au CSV. Crée le fichier avec headers si nécessaire."""
    file_exists = os.path.exists(filepath)
    with open(filepath, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(headers)
        writer.writerow(row)


def _update_excel():
    """Reconstruit le fichier Excel à partir des CSV existants."""
    wb = openpyxl.Workbook()
    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    sheets_config = [
        ("Transactions", CSV_TRANSACTIONS),
        ("Portefeuille", CSV_PORTFOLIO),
        ("Journal", CSV_JOURNAL),
    ]

    for sheet_name, csv_path in sheets_config:
        ws = wb.create_sheet(title=sheet_name)
        if os.path.exists(csv_path):
            with open(csv_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        # Header en gras
                        if row_idx == 1:
                            cell.font = openpyxl.styles.Font(bold=True)
            # Auto-ajuster la largeur des colonnes
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15

    wb.save(EXCEL_FILE)


def log_transaction(transaction: dict, reasoning: str, log_date: date = None):
    """Ajoute une transaction au CSV et met à jour l'Excel.

    `log_date` permet de dater l'entrée sur un jour passé (rattrapage) plutôt
    qu'aujourd'hui ; l'heure reste celle de l'exécution réelle.
    """
    try:
        now = datetime.now()
        row = [
            (log_date or now.date()).strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            transaction.get("action", ""),
            transaction.get("ticker", ""),
            transaction.get("shares", ""),
            transaction.get("price", ""),
            transaction.get("amount_eur", ""),
            transaction.get("fee", ""),
            transaction.get("pnl", ""),
            transaction.get("cash_remaining", ""),
            reasoning,
        ]
        _append_csv(CSV_TRANSACTIONS, HEADERS_TRANSACTIONS, row)
        _update_excel()
        print(f"  ✓ Transaction loguée: {transaction['action']} {transaction['ticker']}")
    except Exception as e:
        print(f"  ⚠ Erreur stockage local (transaction): {e}")


def log_portfolio_summary(portfolio_state: dict, ai_analysis: dict, log_date: date = None):
    """Ajoute un snapshot portefeuille au CSV et met à jour l'Excel.

    `log_date` permet de dater l'entrée sur un jour passé (rattrapage).
    """
    try:
        cash = portfolio_state.get("cash", 0)
        total_invested = portfolio_state.get("total_invested", 0)
        total_value = portfolio_state.get("total_value", 0)
        positions_value = total_value - cash
        pnl = total_value - total_invested
        pnl_pct = (pnl / total_invested * 100) if total_invested > 0 else 0

        row = [
            (log_date or date.today()).strftime("%Y-%m-%d"),
            round(total_value, 2),
            round(cash, 2),
            round(positions_value, 2),
            round(total_invested, 2),
            round(pnl, 2),
            round(pnl_pct, 2),
            portfolio_state.get("nb_positions", 0),
            round(portfolio_state.get("total_fees", 0), 2),
            ai_analysis.get("overall_sentiment", "N/A"),
            ai_analysis.get("risk_level", "N/A"),
        ]
        _append_csv(CSV_PORTFOLIO, HEADERS_PORTFOLIO, row)
        _update_excel()
        print(f"  ✓ Récapitulatif portefeuille logué")
    except Exception as e:
        print(f"  ⚠ Erreur stockage local (portefeuille): {e}")


def log_daily_journal(ai_analysis: dict, transactions_today: list[dict], log_date: date = None):
    """Ajoute l'analyse quotidienne au CSV et met à jour l'Excel.

    `log_date` permet de dater l'entrée sur un jour passé (rattrapage).
    """
    try:
        decisions_text = ""
        for t in transactions_today:
            decisions_text += f"{t['action']} {t['ticker']} ({t.get('amount_eur', 0):.0f}€) | "

        if not decisions_text:
            decisions_text = "HOLD — Aucune action"

        row = [
            (log_date or date.today()).strftime("%Y-%m-%d"),
            ai_analysis.get("market_analysis", "")[:500],
            ai_analysis.get("overall_sentiment", "N/A"),
            ai_analysis.get("risk_level", "N/A"),
            len(transactions_today),
            decisions_text.rstrip(" | "),
        ]
        _append_csv(CSV_JOURNAL, HEADERS_JOURNAL, row)
        _update_excel()
        print(f"  ✓ Journal quotidien logué")
    except Exception as e:
        print(f"  ⚠ Erreur stockage local (journal): {e}")


def setup_local_storage():
    """Initialise les fichiers CSV avec les headers (si non existants)."""
    files = [
        (CSV_TRANSACTIONS, HEADERS_TRANSACTIONS),
        (CSV_PORTFOLIO, HEADERS_PORTFOLIO),
        (CSV_JOURNAL, HEADERS_JOURNAL),
    ]
    for filepath, headers in files:
        if not os.path.exists(filepath):
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow(headers)

    _update_excel()
    print("✓ Stockage local initialisé :")
    print(f"  - {CSV_TRANSACTIONS}")
    print(f"  - {CSV_PORTFOLIO}")
    print(f"  - {CSV_JOURNAL}")
    print(f"  - {EXCEL_FILE}")
    return True
